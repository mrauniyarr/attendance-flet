import flet as ft
from datetime import datetime
import openpyxl
import os
import shutil


# --- SMART STORAGE LOGIC ---
def get_db_path():
    path = os.getenv("FLET_APP_STORAGE_DATA")
    if not path:
        path = os.getcwd()
    return os.path.join(path, "Master_Attendance.xlsx")


async def main(page: ft.Page):  # ← MUST be async
    page.title = "Attendance System (Universal)"
    page.padding = ft.padding.only(top=50, left=20, right=20, bottom=10)
    page.theme_mode = ft.ThemeMode.LIGHT
    page.window_width = 380
    page.window_height = 800
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
    page.scroll = ft.ScrollMode.ADAPTIVE

    EXCEL_FILE = get_db_path()

    # Share service (official way)
    share_service = ft.Share()

    # ====================== PERCENTAGE COLUMN LOGIC (OVERALL CUMULATIVE) ======================
    def update_percentage_column(ws):
        """Adds/updates the 'Attendance %' column as the LAST column.
        Percentage is OVERALL for the entire row (all days so far), NOT per day.
        Example: if 2 days exist and student has P both days → 100.0"""
        if ws.max_row < 2 or ws.max_column < 2:
            return

        # Collect ONLY real date columns (ignore "Attendance %" header)
        date_cols = []
        for c in range(2, ws.max_column + 1):
            header = str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else ""
            if header and header != "Attendance %":
                date_cols.append(c)

        if not date_cols:
            return

        # Percentage column = next blank column after last date
        perc_col = max(date_cols) + 1

        # Set header
        ws.cell(1, perc_col).value = "Attendance %"

        # Calculate OVERALL % for every student row (based on ALL days)
        total_dates = len(date_cols)
        for r in range(2, ws.max_row + 1):
            roll = str(ws.cell(r, 1).value).strip() if ws.cell(r, 1).value else ""
            if not roll or roll == "None":
                continue

            present = sum(
                1
                for c in date_cols
                if ws.cell(r, c).value and str(ws.cell(r, c).value).upper() == "P"
            )

            perc = round((present / total_dates) * 100, 1) if total_dates > 0 else 0.0
            cell = ws.cell(r, perc_col)
            cell.value = perc
            cell.number_format = "0.0"  # Excel shows nicely as 85.5

    def update_all_percentages(e=None):
        """Recalculates overall % for ALL classes."""
        if not os.path.exists(EXCEL_FILE):
            page.snack_bar = ft.SnackBar(ft.Text("No data yet!"), open=True)
            page.update()
            return
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                update_percentage_column(ws)
            wb.save(EXCEL_FILE)
            page.snack_bar = ft.SnackBar(
                ft.Text("All Overall Attendance % updated!"), open=True
            )
            page.update()
        except Exception as ex:
            page.snack_bar = ft.SnackBar(ft.Text(f"Update error: {ex}"), open=True)
            page.update()

    # --- UI CONTROLS (exactly as you wrote) ---
    roll_display = ft.TextField(
        label="Roll Number",
        read_only=True,
        text_align=ft.TextAlign.CENTER,
        width=250,
        text_size=30,
        border_radius=15,
        border_color="red",
    )
    present_count_text = ft.Text(
        "Total Present: 0", size=14, weight="bold", color="green700"
    )

    selected_class = ft.Dropdown(
        label="Select Class",
        width=300,
        value="UG_1",
        border_color="red",
        options=[
            ft.dropdown.Option("UG_1"),
            ft.dropdown.Option("UG_2"),
            ft.dropdown.Option("UG_3"),
            ft.dropdown.Option("PG_1"),
            ft.dropdown.Option("PG_2"),
        ],
    )

    date_display = ft.TextField(
        label="Selected Date",
        value=datetime.now().strftime("%d-%m-%Y"),
        read_only=True,
        width=300,
        suffix_icon=ft.Icons.CALENDAR_MONTH,
        border_color="red",
    )

    async def handle_date_change(e):
        nonlocal date_display
        date_display.value = date_picker.value.strftime("%d-%m-%Y")
        update_present_count()
        page.update()

    def open_date_picker(e):
        date_picker.open = True
        page.update()

    date_display.on_click = open_date_picker

    date_picker = ft.DatePicker(
        on_change=handle_date_change,
        first_date=datetime(2024, 1, 1),
        last_date=datetime(2030, 12, 31),
    )
    page.overlay.append(date_picker)

    # Statistics inputs
    stats_roll_input = ft.TextField(label="Roll No", width=140, border_radius=10)
    stats_class_dropdown = ft.Dropdown(
        label="Class",
        width=140,
        options=[
            ft.dropdown.Option("UG_1"),
            ft.dropdown.Option("UG_2"),
            ft.dropdown.Option("UG_3"),
            ft.dropdown.Option("PG_1"),
            ft.dropdown.Option("PG_2"),
        ],
    )

    res_held = ft.Text("Classes Held: 0", size=16)
    res_pres = ft.Text("Presence: 0", size=16)
    res_perc = ft.Text("Percentage: 0%", size=22, weight="bold")

    def handle_recalc(e):
        if not stats_roll_input.value or not stats_class_dropdown.value:
            page.snack_bar = ft.SnackBar(ft.Text("Fill both fields!"), open=True)
            page.update()
            return

        h, p, pct = get_stats(stats_class_dropdown.value, stats_roll_input.value)
        res_held.value = f"Classes Held: {h}"
        res_pres.value = f"Presence: {p}"
        res_perc.value = f"Percentage: {pct}%"
        res_perc.color = "green" if pct >= 75 else "red"
        page.update()

    # --- CORE LOGIC ---
    def update_present_count(e=None):
        count = 0
        if os.path.exists(EXCEL_FILE):
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
                if selected_class.value in wb.sheetnames:
                    ws = wb[selected_class.value]
                    current_date = date_display.value
                    date_col = None
                    for c in range(2, ws.max_column + 1):
                        if str(ws.cell(row=1, column=c).value) == current_date:
                            date_col = c
                            break
                    if date_col:
                        for r in range(2, ws.max_row + 1):
                            val = ws.cell(row=r, column=date_col).value
                            if val and str(val).upper() == "P":
                                count += 1
                wb.close()
            except:
                pass
        present_count_text.value = f"Total Present: {count}"
        page.update()

    def get_stats(target_class, target_roll):
        """Uses only date columns (ignores Attendance % column)"""
        if not os.path.exists(EXCEL_FILE):
            return 0, 0, 0
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
            if target_class not in wb.sheetnames:
                wb.close()
                return 0, 0, 0
            ws = wb[target_class]

            date_cols = []
            for c in range(2, ws.max_column + 1):
                header = str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else ""
                if header and header != "Attendance %":
                    date_cols.append(c)

            total_classes = len(date_cols)
            if total_classes <= 0:
                wb.close()
                return 0, 0, 0

            target_roll = str(target_roll).strip()
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(r, 1).value).strip() == target_roll:
                    presents = sum(
                        1
                        for c in date_cols
                        if ws.cell(r, c).value
                        and str(ws.cell(r, c).value).upper() == "P"
                    )
                    percent = (
                        round((presents / total_classes) * 100, 1)
                        if total_classes > 0
                        else 0
                    )
                    wb.close()
                    return total_classes, presents, percent
            wb.close()
            return total_classes, 0, 0
        except:
            return 0, 0, 0

    def modify_attendance(action):
        if not roll_display.value:
            page.snack_bar = ft.SnackBar(ft.Text("Enter roll number!"), open=True)
            page.update()
            return

        target_roll = str(roll_display.value).strip()
        class_sheet = selected_class.value
        current_date = date_display.value

        try:
            wb = (
                openpyxl.load_workbook(EXCEL_FILE)
                if os.path.exists(EXCEL_FILE)
                else openpyxl.Workbook()
            )
            ws = (
                wb[class_sheet]
                if class_sheet in wb.sheetnames
                else wb.create_sheet(class_sheet)
            )
            if ws.cell(1, 1).value is None:
                ws.cell(1, 1).value = "Roll / Date"

            date_col = None
            for c in range(2, ws.max_column + 2):
                if str(ws.cell(row=1, column=c).value) == current_date:
                    date_col = c
                    break
                if ws.cell(row=1, column=c).value is None:
                    ws.cell(row=1, column=c).value = current_date
                    date_col = c
                    break

            all_rows = []
            for r in range(2, ws.max_row + 1):
                row_val = [
                    ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)
                ]
                if row_val[0] is not None:
                    all_rows.append(row_val)

            found = False
            for row in all_rows:
                if str(row[0]).strip() == target_roll:
                    row[date_col - 1] = "P" if action == "mark" else ""
                    found = True
                    break
            if not found:
                new_row = [None] * ws.max_column
                new_row[0] = target_roll
                new_row[date_col - 1] = "P" if action == "mark" else ""
                all_rows.append(new_row)

            all_rows.sort(
                key=lambda x: int(str(x[0]).strip()) if str(x[0]).isdigit() else 999
            )
            ws.delete_rows(2, ws.max_row + 10)
            for r_idx, values in enumerate(all_rows, start=2):
                for c_idx, val in enumerate(values, start=1):
                    ws.cell(row=r_idx, column=c_idx).value = val

            wb.save(EXCEL_FILE)

            # Update overall percentage column after every change
            update_percentage_column(ws)

            update_present_count()
            page.snack_bar = ft.SnackBar(
                ft.Text(f"Roll {target_roll} Updated!"), open=True
            )
        except Exception as ex:
            page.snack_bar = ft.SnackBar(ft.Text(f"Error: {ex}"), open=True)

        roll_display.value = ""
        page.update()

    # --- FIXED EXPORT ---
    async def handle_export(e):
        if not os.path.exists(EXCEL_FILE):
            page.snack_bar = ft.SnackBar(ft.Text("No data to export!"), open=True)
            page.update()
            return

        try:
            if page.platform in (ft.PagePlatform.WINDOWS, ft.PagePlatform.LINUX):
                dest = os.path.join(
                    os.path.expanduser("\~"),
                    f"Attendance_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                )
                shutil.copy2(EXCEL_FILE, dest)
                msg = "Saved to Downloads!"
                color = "blue"
            else:  # Android / iOS
                sp = ft.StoragePaths()
                temp_dir = await sp.get_temporary_directory()

                export_name = (
                    f"Attendance_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                temp_file = os.path.join(temp_dir, export_name)

                shutil.copy2(EXCEL_FILE, temp_file)

                result = await share_service.share_files(
                    [ft.ShareFile.from_path(temp_file)], text="Master Attendance Report"
                )
                msg = f"Share sheet opened! ({result.status})"
                color = "green"

            page.snack_bar = ft.SnackBar(ft.Text(msg), open=True, bgcolor=color)
        except Exception as ex:
            page.snack_bar = ft.SnackBar(
                ft.Text(f"Share error: {str(ex)[:80]}"), open=True, bgcolor="red"
            )
        page.update()

    # --- RESET ALL DATA LOGIC ---
    async def handle_reset_confirmed(e):
        confirm_dialog.open = False
        if os.path.exists(EXCEL_FILE):
            try:
                os.remove(EXCEL_FILE)
                update_present_count()
                page.snack_bar = ft.SnackBar(
                    ft.Text("Database Deleted!"), bgcolor="red", open=True
                )
            except Exception as ex:
                page.snack_bar = ft.SnackBar(ft.Text(f"Error: {ex}"), open=True)
        page.update()

    confirm_dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text("Confirm Reset"),
        content=ft.Text(
            "This will permanently delete the Excel file and all attendance data."
        ),
        actions=[
            ft.TextButton(
                "Yes, Delete",
                on_click=handle_reset_confirmed,
            ),
            ft.TextButton(
                "Cancel",
                on_click=lambda _: (
                    setattr(confirm_dialog, "open", False) or page.update()
                ),
            ),
        ],
    )
    page.overlay.append(confirm_dialog)

    def show_reset_dialog(e):
        confirm_dialog.open = True
        page.update()

    # --- KEYPAD ---
    def keypad_btn(val):
        return ft.FilledButton(
            content=ft.Text(str(val), size=20),
            width=75,
            height=75,
            on_click=lambda _: (
                setattr(roll_display, "value", roll_display.value + str(val))
                or page.update()
            ),
        )

    # --- UI LAYOUT (your exact format preserved) ---
    setup_view = ft.Column(
        [
            ft.Text("Configration:", size=20, weight="bold", color="blue"),
            ft.Container(height=10),
            date_display,
            ft.Container(height=15),
            selected_class,
            ft.Container(height=15),
            ft.FilledButton(
                "Share_Excel",
                icon=ft.Icons.SHARE,
                on_click=handle_export,
                width=300,
            ),
            ft.Container(height=15),
            ft.FilledButton(
                "Reset_All",
                icon=ft.Icons.DELETE_FOREVER,
                on_click=show_reset_dialog,
                width=300,
                bgcolor="blue",
            ),
            ft.Container(height=10),
            ft.FilledButton(
                "Refresh All %",
                icon=ft.Icons.REFRESH,
                on_click=update_all_percentages,
                width=300,
            ),
            present_count_text,
            ft.Divider(height=40, thickness=1),
            ft.Container(
                content=ft.Column(
                    [
                        ft.Text(
                            "Developer: Dr Manish kumar Gupta", size=12, weight="w500"
                        ),
                        ft.Text("Asst. Professor,Faculty of Commerce", size=10),
                        ft.Text(
                            "SMM Town PG College, Ballia",
                            size=10,
                            italic=True,
                            color="w500",
                        ),
                    ],
                    spacing=5,
                    horizontal_alignment="center",
                ),
                padding=20,
                border=ft.border.all(1, ft.Colors.BLUE_GREY_100),
                border_radius=15,
                width=300,
            ),
            ft.Text("Version 1.0.0", size=4, color="grey400"),
        ],
        horizontal_alignment="center",
    )

    statistics_view = ft.Column(
        [
            ft.Container(height=20),
            ft.Text(
                "STUDENT STATISTICS", size=22, weight="bold", color="blue_grey_700"
            ),
            ft.Row(
                [
                    stats_class_dropdown,
                    stats_roll_input,
                ],
                alignment="center",
                spacing=10,
            ),
            ft.Container(height=15),
            ft.ElevatedButton(
                "RECALCULATE",
                icon=ft.Icons.REFRESH,
                on_click=handle_recalc,
                width=290,
                icon_color="blue",
                color="red",
                bgcolor="blue_700",
                elevation=10,
            ),
            ft.Container(height=20),
            ft.Container(
                content=ft.Column(
                    [res_held, res_pres, ft.Divider(thickness=1), res_perc],
                    horizontal_alignment="center",
                    spacing=12,
                ),
                padding=25,
                bgcolor=ft.Colors.GREY_50,
                border_radius=25,
                width=320,
                border=ft.Border.all(1, ft.Colors.BLUE_GREY_100),
            ),
        ],
        horizontal_alignment="center",
        visible=False,
    )

    attendance_view = ft.Column(
        [
            ft.Text("MARKING", size=20, weight="bold"),
            roll_display,
            ft.Row([keypad_btn(1), keypad_btn(2), keypad_btn(3)], alignment="center"),
            ft.Row([keypad_btn(4), keypad_btn(5), keypad_btn(6)], alignment="center"),
            ft.Row([keypad_btn(7), keypad_btn(8), keypad_btn(9)], alignment="center"),
            ft.Row(
                [
                    ft.IconButton(
                        ft.Icons.DELETE,
                        on_click=lambda _: (
                            setattr(roll_display, "value", "") or page.update()
                        ),
                    ),
                    keypad_btn(0),
                    ft.IconButton(
                        ft.Icons.BACKSPACE,
                        on_click=lambda _: (
                            setattr(roll_display, "value", roll_display.value[:-1])
                            or page.update()
                        ),
                    ),
                ],
                alignment="center",
            ),
            ft.Row(
                [
                    ft.FilledButton(
                        "Save",
                        on_click=lambda _: modify_attendance("mark"),
                        width=140,
                        bgcolor="green",
                    ),
                    ft.FilledButton(
                        "Remove",
                        on_click=lambda _: modify_attendance("remove"),
                        width=140,
                        bgcolor="red",
                    ),
                ],
                alignment="center",
            ),
        ],
        horizontal_alignment="center",
        visible=False,
    )

    def switch_nav(e):
        setup_view.visible = e.control.data == "s"
        attendance_view.visible = e.control.data == "m"
        statistics_view.visible = e.control.data == "stat"
        page.update()

    page.add(
        ft.Row(
            [
                ft.TextButton("SETUP", data="s", on_click=switch_nav),
                ft.TextButton("MARKING", data="m", on_click=switch_nav),
                ft.TextButton("STATISTICS", data="stat", on_click=switch_nav),
            ],
            alignment="center",
        ),
        ft.Divider(),
        setup_view,
        attendance_view,
        statistics_view,
    )

    update_present_count()
    update_all_percentages()  # Creates the overall % column on first run


# --- RUN THE APP ---
ft.run(main)

